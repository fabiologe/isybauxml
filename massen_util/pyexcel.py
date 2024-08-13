import openpyxl as xl
import pandas as pd
import os 
col_bauwerk = {
    "Status":{
        "start_cell": "A5",
        "data_type" : "float"
    },
    "Bauwerk":{
        "start_cell": "B5",
        "data_type": "str",
    }, 
    "Bauwerkart":{
        "start_cell": "C5",
        "data_type":"str",
    },
    "Tiefe": {
        "start_cell": "D5",
        "data_type": "float",
    },
    "Breite OK":{
        "start_cell": "E5",
        "data_type": "float"
    },
    "Laenge OK":{
        "start_cell": "F5",
        "data_type": "float"
    },
    "Flaeche OK":{
        "start_cell": "G5",
        "data_type": "float",
    },
    "Flaeche UK":{
        "start_cell": "H5",
        "data_type": "float",
    },
    "Volumen 1":{
        "start_cell": "I5",
        "data_type": "float",
    }, 
    "Volumen 2":{
        "start_cell": "J5",
        "data_type": "float"
    }
}
col_leitung = {
    "DN":{
        "start_cell": "A6",
        "data_type": "float"
    },
    "Laenge":{
        "start_cell": "F6",
        "data_type": "float"
    }
}
col_schacht = {
    "Schacht": {
        "start_cell":"A5",
        "data_type":"str"
    },
    "Tiefe": {
        "start_cell": "B5",
        "data_type":"float"
    },
    "DN Zulauf":{
        "start_cell": "C5",
        "data_type": "float"
    },
    "DN Ablauf":{
        "start_cell":"D5", 
        "data_type": "float"
    }
}
col_haltung = {
    "Status": {
        "start_cell": "W7",
        "data_type": "str"
    },
    "Knoten Nr. oben": {
        "start_cell": "A7",
        "data_type": "str"
    },
    "Knoten Nr. unten": {
        "start_cell": "B7",
        "data_type": "str"
    },
    "Deckelhoehe oben": {
        "start_cell": "C7",
        "data_type": "float"
    },
    "Deckelhoehe unten": {
        "start_cell": "D7",
        "data_type": "float"
    },
    "Sohlhoehe oben": {
        "start_cell": "E7",
        "data_type": "float"
    },
    "Sohlhoehe unten": {
        "start_cell": "F7",
        "data_type": "float"
    },
    "Laenge": {
        "start_cell": "G7",
        "data_type": "float"
    },
    "DN": {
        "start_cell": "L7",
        "data_type": "float"
    }
}
def to_xsls_haltung(col_haltung, col_schacht, col_leitung, col_bauwerk):
    csv_h = os.path.abspath('storage/output_xlsx_csv/massen_haltungen.csv')
    csv_s = os.path.abspath('storage/output_xlsx_csv/massen_schacht.csv')
    csv_l = os.path.abspath('storage/output_xlsx_csv/massen_leitung.csv')
    csv_b = os.path.abspath('storage/output_xlsx_csv/massen_bauwerk.csv')
    src = os.path.abspath('massen_util/src/massen.xlsx')
   
    wb = xl.load_workbook(src)
    ws_haltung = wb["Massen_Haltung"]
    ws_leitung = wb["Massen_Leitung"]
    ws_schacht = wb["Massen_Schacht"]
    ws_bauwerk = wb["Massen_Bauwerk"]

    df_s = pd.read_csv(csv_s)
    df_h = pd.read_csv(csv_h)
    df_b = pd.read_csv(csv_b)
    # Haltung-CSV to xsls
    
    for col_name, assignment in col_schacht.items():
        data_list = df_s[col_name].tolist()
        start_row, start_col = xl.utils.coordinate_to_tuple(assignment["start_cell"])
        start_row = max(start_row, 1)
        start_col = max(start_col, 1)

        for row_idx, value in enumerate(data_list):
            # Apply correct data type formatting before writing
            if assignment["data_type"] == "str":
                cell_value = str(value)
            else:
                cell_value = float(value)

            # Write to the cell based on row and column offset
            ws_schacht.cell(row=row_idx + start_row, column=start_col).value = cell_value

    for col_name, assignment in col_haltung.items():
        data_list = df_h[col_name].tolist()  

       
        start_row, start_col = xl.utils.coordinate_to_tuple(assignment["start_cell"])

       
        start_row = max(start_row, 1)
        start_col = max(start_col, 1)

        for row_idx, value in enumerate(data_list):
            # Apply correct data type formatting before writing
            if assignment["data_type"] == "str":
                cell_value = str(value)
            else:
                cell_value = float(value)

            # Write to the cell based on row and column offset
            ws_haltung.cell(row=row_idx + start_row, column=start_col).value = cell_value

    for col_name, assignment in col_bauwerk.items():
        data_list = df_b[col_name].tolist()
        start_row, start_col = xl.utils.coordinate_to_tuple(assignment["start_cell"])
        start_row = max(start_row, 1)
        start_col = max(start_col, 1)

        for row_idx, value in enumerate(data_list):
            # Apply correct data type formatting before writing
            if assignment["data_type"] == "str":
                cell_value = str(value)
            else:
                cell_value = float(value)

            # Write to the cell based on row and column offset
            ws_bauwerk.cell(row=row_idx + start_row, column=start_col).value = cell_value

    wb.save('storage/output_xlsx_csv/massen_gesamt.xlsx')
    return print("Data successfully written to your XLS file!")

